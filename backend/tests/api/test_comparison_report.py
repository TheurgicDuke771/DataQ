"""Comparison report download tests (ADR 0015 §4, #795) — real Postgres.

The report must be derived from the REDACTED buckets (policy applied exactly
like the run-detail read), 404 on unknown run/result, 422 on a non-comparison
result, and never persist anything server-side. Skips without TEST_DATABASE_URL.
"""

import io
import uuid
from collections.abc import Iterator
from typing import Any

import pytest
from fastapi.testclient import TestClient

from backend.app.db.models import Check, Connection, Result, Run, Suite, User
from backend.app.db.session import get_db
from backend.app.main import app


@pytest.fixture
def client(db_session: Any) -> Iterator[TestClient]:
    app.dependency_overrides[get_db] = lambda: db_session
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


def _fixture_run(
    client: TestClient, db_session: Any, *, kind: str = "comparison"
) -> tuple[Run, Result]:
    # Connections via ORM (no validation needed); the SUITE via the API so it
    # is owned by the dev-bypass identity `require_permission` resolves.
    owner = User(aad_object_id=uuid.uuid4().hex, email=f"{uuid.uuid4().hex[:8]}@example.com")
    db_session.add(owner)
    db_session.flush()
    conn = Connection(
        name=f"sf-{uuid.uuid4().hex[:8]}",
        type="snowflake",
        env="dev",
        config={},
        created_by=owner.id,
    )
    source = Connection(
        name=f"src-{uuid.uuid4().hex[:8]}",
        type="snowflake",
        env="dev",
        config={},
        created_by=owner.id,
    )
    db_session.add_all([conn, source])
    db_session.commit()
    created = client.post(
        "/api/v1/suites",
        json={"name": f"recon-{uuid.uuid4().hex[:6]}", "connection_id": str(conn.id)},
    )
    assert created.status_code == 201
    suite = db_session.get(Suite, uuid.UUID(created.json()["id"]))
    suite.column_policy = {"identifier_column": "order_id", "pii_columns": ["email"]}
    db_session.flush()
    check = Check(
        suite_id=suite.id,
        name="orders reconcile",
        kind=kind,
        expectation_type="comparison:records" if kind == "comparison" else "x",
        source_connection_id=source.id if kind == "comparison" else None,
        config={"source": {"table": "T"}, "keys": ["order_id"]} if kind == "comparison" else {},
    )
    db_session.add(check)
    db_session.flush()
    run = Run(suite_id=suite.id, status="succeeded")
    db_session.add(run)
    db_session.flush()
    result = Result(
        run_id=run.id,
        check_id=check.id,
        status="fail",
        metric_value=50,
        observed_value={
            "source_rows": 2,
            "target_rows": 2,
            "matched": 1,
            "mismatched": 1,
            "additional_in_source": 0,
            "additional_in_target": 0,
            "mismatch_percent": 50.0,
        },
        sample_failures={
            "mismatched": [{"order_id": "7", "email_src": "a@x.io", "email_tgt": "b@x.io"}]
        },
    )
    db_session.add(result)
    db_session.commit()
    return run, result


def test_csv_report_is_redacted_and_carries_buckets(client: TestClient, db_session: Any) -> None:
    run, result = _fixture_run(client, db_session)
    resp = client.get(f"/api/v1/runs/{run.id}/results/{result.id}/comparison_report?fmt=csv")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "attachment" in resp.headers["content-disposition"]
    body = resp.text
    assert "# mismatch_percent,50.0" in body
    assert "mismatched,7," in body  # identifier shown, bucket labeled
    # PII values must never reach the file — redaction runs before formatting.
    assert "a@x.io" not in body and "b@x.io" not in body
    assert "<redacted>" in body


def test_xlsx_report_has_workbook_magic_and_sheets(client: TestClient, db_session: Any) -> None:
    run, result = _fixture_run(client, db_session)
    resp = client.get(f"/api/v1/runs/{run.id}/results/{result.id}/comparison_report?fmt=xlsx")
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"  # zip container
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["summary", "mismatched"]
    mismatched = wb["mismatched"]
    cells = [str(c.value) for row in mismatched.iter_rows() for c in row]
    assert "a@x.io" not in cells and "<redacted>" in cells


def test_non_comparison_result_is_422(client: TestClient, db_session: Any) -> None:
    run, result = _fixture_run(client, db_session, kind="expectation")
    resp = client.get(f"/api/v1/runs/{run.id}/results/{result.id}/comparison_report")
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "comparison_report_invalid"


def test_unknown_run_and_result_are_404(client: TestClient, db_session: Any) -> None:
    run, _ = _fixture_run(client, db_session)
    missing = client.get(f"/api/v1/runs/{uuid.uuid4()}/results/{uuid.uuid4()}/comparison_report")
    assert missing.status_code == 404
    wrong_result = client.get(f"/api/v1/runs/{run.id}/results/{uuid.uuid4()}/comparison_report")
    assert wrong_result.status_code == 404


def test_bad_format_is_422(client: TestClient, db_session: Any) -> None:
    run, result = _fixture_run(client, db_session)
    resp = client.get(f"/api/v1/runs/{run.id}/results/{result.id}/comparison_report?fmt=pdf")
    assert resp.status_code == 422
